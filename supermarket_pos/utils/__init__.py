# ============================================================
# utils/__init__.py — Utilidades del sistema
# ============================================================

from __future__ import annotations
import re
import os
from datetime import datetime
from typing import Optional


# ============================================================
# VALIDATORS
# ============================================================
class Validators:
    EMAIL_RE = re.compile(r"^[\w.+-]+@[\w-]+\.[\w.-]+$")
    PHONE_RE = re.compile(r"^\+?[\d\s\-]{7,15}$")
    ID_RE    = re.compile(r"^[\d\-\.]{5,20}$")

    @staticmethod
    def email(value: str) -> bool:
        return bool(Validators.EMAIL_RE.match(value.strip())) if value else True

    @staticmethod
    def phone(value: str) -> bool:
        return bool(Validators.PHONE_RE.match(value.strip())) if value else True

    @staticmethod
    def identification(value: str) -> bool:
        return bool(Validators.ID_RE.match(value.strip())) if value else False

    @staticmethod
    def positive_number(value) -> bool:
        try:
            return float(value) >= 0
        except (ValueError, TypeError):
            return False

    @staticmethod
    def positive_int(value) -> bool:
        try:
            return int(value) >= 0
        except (ValueError, TypeError):
            return False

    @staticmethod
    def not_empty(value: str) -> bool:
        return bool(str(value).strip())


# ============================================================
# FORMATTERS
# ============================================================
class Formatters:
    @staticmethod
    def currency(value: float) -> str:
        return f"$ {value:,.0f}"

    @staticmethod
    def date(dt: Optional[datetime]) -> str:
        if dt is None:
            return ""
        return dt.strftime("%d/%m/%Y")

    @staticmethod
    def datetime(dt: Optional[datetime]) -> str:
        if dt is None:
            return ""
        return dt.strftime("%d/%m/%Y %H:%M:%S")

    @staticmethod
    def percent(value: float) -> str:
        return f"{value:.2f}%"


# ============================================================
# PDF INVOICE GENERATOR (ReportLab)
# ============================================================
class InvoicePDFGenerator:
    """Genera facturas en PDF usando ReportLab."""

    def __init__(self):
        self._available = self._check_reportlab()

    @staticmethod
    def _check_reportlab() -> bool:
        try:
            import reportlab
            return True
        except ImportError:
            return False

    def generate(self, sale, output_path: str) -> str:
        """
        Genera el PDF de la factura.
        Returns: ruta del archivo generado.
        """
        if not self._available:
            raise RuntimeError(
                "ReportLab no está instalado.\n"
                "Ejecuta: pip install reportlab --break-system-packages"
            )
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, Image as RLImage, HRFlowable
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
        from config import (COMPANY_NAME, COMPANY_NIT, COMPANY_ADDRESS,
                            COMPANY_PHONE, COMPANY_EMAIL, LOGO_PATH,
                            THEME_PRIMARY, THEME_SECONDARY)

        styles = getSampleStyleSheet()
        primary = colors.HexColor(THEME_PRIMARY)
        gold    = colors.HexColor(THEME_SECONDARY)

        title_style = ParagraphStyle("title", parent=styles["Normal"],
                                     fontSize=14, textColor=primary,
                                     alignment=TA_CENTER, fontName="Helvetica-Bold")
        sub_style   = ParagraphStyle("sub",   parent=styles["Normal"],
                                     fontSize=9,  textColor=colors.grey,
                                     alignment=TA_CENTER)
        label_style = ParagraphStyle("label", parent=styles["Normal"],
                                     fontSize=9,  fontName="Helvetica-Bold")
        value_style = ParagraphStyle("value", parent=styles["Normal"], fontSize=9)
        total_style = ParagraphStyle("total", parent=styles["Normal"],
                                     fontSize=12, fontName="Helvetica-Bold",
                                     textColor=primary, alignment=TA_RIGHT)

        doc  = SimpleDocTemplate(output_path, pagesize=A4,
                                 topMargin=1.5*cm, bottomMargin=1.5*cm,
                                 leftMargin=2*cm, rightMargin=2*cm)
        elems = []

        # --- Logo + Encabezado ---
        header_data = []
        if os.path.exists(LOGO_PATH):
            logo = RLImage(LOGO_PATH, width=2.5*cm, height=2.5*cm)
            header_data.append([logo,
                                 [Paragraph(COMPANY_NAME, title_style),
                                  Paragraph(f"NIT: {COMPANY_NIT}", sub_style),
                                  Paragraph(COMPANY_ADDRESS, sub_style),
                                  Paragraph(f"Tel: {COMPANY_PHONE}  |  {COMPANY_EMAIL}", sub_style)]])
        else:
            header_data.append(["",
                                 [Paragraph(COMPANY_NAME, title_style),
                                  Paragraph(f"NIT: {COMPANY_NIT}", sub_style)]])

        t = Table(header_data, colWidths=[3*cm, None])
        t.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "MIDDLE")]))
        elems.append(t)
        elems.append(Spacer(1, 0.3*cm))
        elems.append(HRFlowable(width="100%", thickness=2, color=primary))
        elems.append(Spacer(1, 0.3*cm))

        # --- Datos de factura ---
        sold_str = sale.sold_at.strftime("%d/%m/%Y %H:%M:%S") if sale.sold_at else ""
        info_data = [
            [Paragraph("<b>FACTURA DE VENTA</b>", label_style),
             Paragraph(f"<b>N°:</b> {sale.invoice_number}", label_style)],
            [Paragraph(f"<b>Fecha:</b> {sold_str}", value_style),
             Paragraph(f"<b>Cajero:</b> {sale.employee_name}", value_style)],
            [Paragraph(f"<b>Cliente:</b> {sale.customer_name}", value_style),
             Paragraph(f"<b>Pago:</b> {sale.payment_method.capitalize()}", value_style)],
        ]
        t2 = Table(info_data, colWidths=[None, None])
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#e8f5e9")),
            ("GRID", (0,0), (-1,-1), 0.5, colors.lightgrey),
            ("PADDING", (0,0), (-1,-1), 4),
        ]))
        elems.append(t2)
        elems.append(Spacer(1, 0.3*cm))

        # --- Tabla de productos ---
        prod_headers = [
            Paragraph("<b>#</b>",         label_style),
            Paragraph("<b>Descripción</b>", label_style),
            Paragraph("<b>Cant.</b>",     label_style),
            Paragraph("<b>P. Unit.</b>",  label_style),
            Paragraph("<b>Total</b>",     label_style),
        ]
        rows = [prod_headers]
        for i, d in enumerate(sale.details, 1):
            rows.append([
                str(i),
                d.product_name,
                str(d.quantity),
                f"$ {d.unit_price:,.0f}",
                f"$ {d.subtotal:,.0f}",
            ])
        pt = Table(rows, colWidths=[1*cm, None, 1.8*cm, 3*cm, 3*cm])
        pt.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), primary),
            ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f5f5f5")]),
            ("GRID",   (0,0), (-1,-1), 0.5, colors.lightgrey),
            ("PADDING",(0,0), (-1,-1), 5),
            ("ALIGN",  (2,0), (-1,-1), "RIGHT"),
        ]))
        elems.append(pt)
        elems.append(Spacer(1, 0.4*cm))

        # --- Totales ---
        tot_data = [
            ["",  "Subtotal:",    f"$ {sale.subtotal:,.0f}"],
            ["",  "IVA (19%):",   f"$ {sale.tax_amount:,.0f}"],
            ["",  "TOTAL:",       f"$ {sale.total:,.0f}"],
            ["",  "Pagado:",      f"$ {sale.paid_amount:,.0f}"],
            ["",  "Cambio:",      f"$ {sale.change_amount:,.0f}"],
        ]
        tt = Table(tot_data, colWidths=[None, 4*cm, 4*cm])
        tt.setStyle(TableStyle([
            ("ALIGN",  (1,0), (-1,-1), "RIGHT"),
            ("FONTNAME",(1,2), (-1,2), "Helvetica-Bold"),
            ("FONTSIZE",(1,2), (-1,2), 12),
            ("TEXTCOLOR",(1,2),(-1,2), primary),
            ("LINEABOVE",(1,2),(-1,2), 1, primary),
            ("PADDING",(0,0),(-1,-1), 4),
        ]))
        elems.append(tt)
        elems.append(Spacer(1, 0.5*cm))
        elems.append(HRFlowable(width="100%", thickness=1, color=colors.lightgrey))
        elems.append(Spacer(1, 0.2*cm))
        elems.append(Paragraph("¡Gracias por su compra! — www.dto2.com", sub_style))

        doc.build(elems)
        return output_path


# ============================================================
# EXCEL REPORT EXPORTER
# ============================================================
class ExcelExporter:
    """Exporta reportes a Excel usando openpyxl."""

    @staticmethod
    def export(data: list[dict], filepath: str, sheet_name: str = "Reporte") -> str:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise RuntimeError(
                "openpyxl no está instalado.\n"
                "Ejecuta: pip install openpyxl --break-system-packages"
            )
        if not data:
            raise ValueError("No hay datos para exportar.")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        headers = list(data[0].keys())
        header_fill = PatternFill("solid", fgColor="1a6b36")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin")
        )

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=str(h).upper())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin

        for row_idx, record in enumerate(data, 2):
            fill = PatternFill("solid", fgColor="f5f5f5") if row_idx % 2 == 0 else None
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col, value=record.get(h))
                cell.border = thin
                if fill:
                    cell.fill = fill

        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        wb.save(filepath)
        return filepath
